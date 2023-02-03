from openpyxl import load_workbook
import datetime

file = open("mebros.txt", "w")
wb = load_workbook(filename = 'data.xlsx')
string = ''
length = 0

wb = wb['Membros 2021']
months = {1 : 'janeiro', 2 : 'fevereiro', 3 : 'março', 4 : 'abril', 5 : 'maio', 6 : 'junho', 7 : 'julho', 8 : 'agosto', 9 : 'setembro', 10 : 'outubro', 11 : 'novembro', 12 : 'dezembro'}

def date_convert(date):
    if (type(date) is datetime.datetime):
        date = (months[date.month] + ' de ' + str(date.year))
    return date

def to_lower(string):
    if string:
        if string[-1] == '\n':
            string = string[:-1]
        return string.lower()
    return string

for row in wb.values:
    if row[0] != None:
        length += 1

for row in wb.iter_rows(min_row = 2, max_col = 12, max_row = length, values_only = True):
        string += ( '{}, {}, {}, {}, portador(a) do RG nº. {} e CPF sob o n°. {}, residente e domiciliado(a) em {}, voluntário(a) na empresa júnior desde {}.\n'.format(row[0],to_lower(row[9]),to_lower(row[10]),to_lower(row[11]),row[5],row[6],row[7],date_convert(row[4])) )
        
file.write(string)


